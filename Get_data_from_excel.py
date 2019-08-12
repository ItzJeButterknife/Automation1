import openpyxl
from openpyxl.styles import PatternFill
data = openpyxl.load_workbook('testfile.xlsx')
sheet = data.get_sheet_by_name('Blad1')
startrow = 2
persons = []
family = False

class Person:
    def __init__(self, forname, surname, adress, date, relation):
        self.forname = forname
        self.surname = surname
        self.adress = adress
        self.date = date
        self.relation = relation

def get_data(row):
    global street, number, adress, surname, forname, date
    street = sheet['C'+str(row)].value
    number = sheet['D'+str(row)].value
    adress = str(street) + " " + str(number)

    surname = sheet['B'+str(row)].value
    forname = sheet['A'+str(row)].value

    date = sheet['F'+str(row)].value

row = startrow

sheet['A1'].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")

while family:
    nextrow = row + 1
    get_data(row)
    p = Person(forname,surname,adress,date, 0)
    persons.append(p)
    adress1 = adress
    get_data(nextrow)

    if adress == adress1:
        row += 1
    else:
        family = False
        startrow = 6

for i in range(1,len(persons)):
    difference = abs(persons[0].date - persons[i].date)
    difference = str(difference / 365)
    location = difference.find(' ')
    difference = int(difference[0:2])
    if difference <= 10:
        persons[i].relation = 1
    elif difference >= 10:
        persons[i].relation = 2
    else:
        persons[i].relation = 0
